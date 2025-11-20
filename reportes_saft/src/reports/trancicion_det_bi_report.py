import pandas as pd
from openpyxl.styles import Alignment, Font, PatternFill, Border, Side
from openpyxl.utils import get_column_letter
from openpyxl.drawing.image import Image as XLImage
from datetime import datetime


class TrancicionBIDetalleReport:
    def __init__(self,  municipio,  municipio_admin=False):
        self.municipio = municipio
        self.muni_admin = municipio_admin

    def generar_excel(self, bi_urbano, bi_rural, bi_urbano_act, bi_rurales_act, bi_tec_urbano_inicio, bi_tec_rural_inicio, bi_tec_urbano_final, bi_tec_rurales_final, bi_dec_rural_inicio,  bi_dec_urbano_inicio, bi_dec_urbano_final, bi_dec_rurales_final, ruta_excel: str, titulo_rpt: str):
        header_fill = PatternFill(
            start_color="1F4E78", end_color="1F4E78", fill_type="solid")
        header_font = Font(color="FFFFFF", bold=True)
        border_style = Border(
            left=Side(border_style="thin"),
            right=Side(border_style="thin"),
            top=Side(border_style="thin"),
            bottom=Side(border_style="thin")
        )

        with pd.ExcelWriter(ruta_excel, engine="openpyxl") as writer:
            datos = {
                f"{titulo_rpt} Urbano": pd.DataFrame(bi_urbano),
                f"{titulo_rpt} Rural": pd.DataFrame(bi_rural),
                f"{titulo_rpt} Urbano Activos": pd.DataFrame(bi_urbano_act),
                f"{titulo_rpt} Rural Activos": pd.DataFrame(bi_rurales_act),

                f"{titulo_rpt} Tecnificado Urbano Inicio": pd.DataFrame(bi_tec_urbano_inicio),
                f"{titulo_rpt} Tecnificado Rural Inicio": pd.DataFrame(bi_tec_rural_inicio),
                f"{titulo_rpt} Tecnificado Urbano Final": pd.DataFrame(bi_tec_urbano_final),
                f"{titulo_rpt} Tecnificado Rural Final": pd.DataFrame(bi_tec_rurales_final),

                f"{titulo_rpt} Declarado Urbano Inicio": pd.DataFrame(bi_dec_urbano_inicio),
                f"{titulo_rpt} Declarado Rural Inicio": pd.DataFrame(bi_dec_rural_inicio),
                f"{titulo_rpt} Declarado Urbano Final": pd.DataFrame(bi_dec_urbano_final),
                f"{titulo_rpt} Declarado Rural Final": pd.DataFrame(bi_dec_rurales_final),

            }

            for nombre_hoja, df in datos.items():
                if len(df) == 0:
                    continue
                df.to_excel(writer, sheet_name=nombre_hoja,
                            index=False, startrow=4)
                ws = writer.sheets[nombre_hoja]

                col_fin = len(df.columns)

                # -------------------------------------------------------
                # 1️⃣ TÍTULO CENTRADO
                # -------------------------------------------------------
                ws.merge_cells(start_row=1, start_column=1,
                               end_row=1, end_column=col_fin)
                celda_titulo = ws.cell(row=1, column=1)
                celda_titulo.value = f"REPORTE {nombre_hoja.upper()}"
                celda_titulo.font = Font(size=15, bold=True)
                celda_titulo.alignment = Alignment(horizontal="center")

                # -------------------------------------------------------
                # 2️⃣ FECHA DE EMISIÓN
                # -------------------------------------------------------
                ws.merge_cells(start_row=2, start_column=1,
                               end_row=2, end_column=col_fin)
                celda_fecha = ws.cell(row=2, column=1)
                celda_fecha.value = f"Fecha de emisión: {datetime.now().strftime('%d/%m/%Y')}"
                celda_fecha.font = Font(size=11)
                celda_fecha.alignment = Alignment(horizontal="left")

                # -------------------------------------------------------
                # 3️⃣ TOTAL DE REGISTROS
                # -------------------------------------------------------
                ws.merge_cells(start_row=3, start_column=1,
                               end_row=3, end_column=col_fin)
                celda_total = ws.cell(row=3, column=1)
                celda_total.value = f"Total de registros: {len(df)}"
                celda_total.font = Font(size=11, bold=True)

                # -------------------------------------------------------
                # 4️⃣ FORMATO DE ENCABEZADOS
                # -------------------------------------------------------
                for col in range(1, col_fin + 1):
                    cell = ws.cell(row=5, column=col)
                    cell.fill = header_fill
                    cell.font = header_font
                    cell.alignment = Alignment(horizontal="center")

                # -------------------------------------------------------
                # 5️⃣ BORDES PARA TODA LA TABLA
                # -------------------------------------------------------
                for row in ws.iter_rows(min_row=5, max_row=5 + len(df), min_col=1, max_col=col_fin):
                    for cell in row:
                        cell.border = border_style

                # -------------------------------------------------------
                # 6️⃣ AUTOAJUSTE DE COLUMNAS
                # -------------------------------------------------------
                for col in ws.columns:
                    max_length = 0
                    col_letter = get_column_letter(col[0].column)

                    for cell in col:
                        if cell.value:
                            max_length = max(max_length, len(str(cell.value)))

                    ws.column_dimensions[col_letter].width = max(
                        max_length + 2, 12)

                # -------------------------------------------------------
                # 7️⃣ FILTROS AUTOMÁTICOS
                # -------------------------------------------------------
                ws.auto_filter.ref = f"A5:{get_column_letter(col_fin)}{4 + len(df)}"

                # -------------------------------------------------------
                # 8️⃣ CONGELAR FILA DE ENCABEZADOS
                # -------------------------------------------------------
                ws.freeze_panes = "A6"

        return ruta_excel
