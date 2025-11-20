import pandas as pd
from openpyxl.styles import Alignment, Font, PatternFill, Border, Side
from openpyxl.utils import get_column_letter
from openpyxl.drawing.image import Image as XLImage
from datetime import datetime


class TrancicionSPDetalleReport:
    def __init__(self,  municipio,  municipio_admin=False):
        self.municipio = municipio
        self.muni_admin = municipio_admin

    def generar_excel(self, resultados, ruta_excel: str, titulo_rpt: str):
        header_fill = PatternFill(
            start_color="1F4E78", end_color="1F4E78", fill_type="solid")
        header_font = Font(color="FFFFFF", bold=True)
        border_style = Border(
            left=Side(border_style="thin"),
            right=Side(border_style="thin"),
            top=Side(border_style="thin"),
            bottom=Side(border_style="thin")
        )
        hoja_creada = False
        with pd.ExcelWriter(ruta_excel, engine="openpyxl") as writer:
            for nombre_servicio, datos in resultados.items():
                for tipo, df in datos.items():

                    if df is None or len(df) == 0:
                        continue
                    hoja_nombre = f"{nombre_servicio} {tipo}"[
                        :30]  # Excel limita 31 caracteres
                    if isinstance(df, list):
                        df = pd.DataFrame(df)
                    df.to_excel(writer, sheet_name=hoja_nombre,
                                index=False, startrow=4)
                    ws = writer.sheets[hoja_nombre]
                    col_fin = len(df.columns)

                    # TÍTULO
                    ws.merge_cells(start_row=1, start_column=1,
                                   end_row=1, end_column=col_fin)
                    celda_titulo = ws.cell(row=1, column=1)
                    celda_titulo.value = f"{nombre_servicio.upper()} - {tipo.upper()}"
                    celda_titulo.font = Font(size=14, bold=True)
                    celda_titulo.alignment = Alignment(horizontal="center")

                    # FECHA DE EMISIÓN
                    ws.merge_cells(start_row=2, start_column=1,
                                   end_row=2, end_column=col_fin)
                    celda_fecha = ws.cell(row=2, column=1)
                    celda_fecha.value = f"Fecha de emisión: {datetime.now().strftime('%d/%m/%Y')}"
                    celda_fecha.alignment = Alignment(horizontal="left")

                    # TOTAL DE REGISTROS
                    ws.merge_cells(start_row=3, start_column=1,
                                   end_row=3, end_column=col_fin)
                    celda_total = ws.cell(row=3, column=1)
                    celda_total.value = f"Total de registros: {len(df)}"
                    celda_total.font = Font(bold=True)

                    # ENCABEZADOS
                    for col in range(1, col_fin + 1):
                        cell = ws.cell(row=4, column=col)
                        cell.fill = header_fill
                        cell.font = header_font
                        cell.alignment = Alignment(horizontal="center")
                        cell.border = border_style

                    # BORDES TABLA
                    for row in ws.iter_rows(min_row=5, max_row=4 + len(df), min_col=1, max_col=col_fin):
                        for cell in row:
                            cell.border = border_style

                    # AUTOAJUSTE COLUMNAS
                    for col in ws.columns:
                        max_length = 0
                        col_letter = get_column_letter(col[0].column)
                        for cell in col:
                            if cell.value:
                                max_length = max(
                                    max_length, len(str(cell.value)))
                        ws.column_dimensions[col_letter].width = max(
                            max_length + 2, 12)

                    # FILTROS AUTOMÁTICOS
                    ws.auto_filter.ref = f"A4:{get_column_letter(col_fin)}{4 + len(df)}"

                    # CONGELAR FILA DE ENCABEZADOS
                    ws.freeze_panes = "A5"

        return ruta_excel
