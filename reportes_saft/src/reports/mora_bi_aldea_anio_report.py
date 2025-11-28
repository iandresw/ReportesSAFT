from decimal import Decimal
from reportlab.lib.pagesizes import letter, landscape
from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer, Image
from reportlab.lib import colors
from reportlab.lib.styles import getSampleStyleSheet
from openpyxl.utils import get_column_letter
from datetime import datetime
import matplotlib.pyplot as plt
from openpyxl import Workbook
from openpyxl.chart import BarChart, Reference, series
from openpyxl.chart.shapes import GraphicalProperties
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from openpyxl.drawing.image import Image as Img
import pandas as pd
from ui.ui_style_table import columa_style


class MoravsBIAldeaAnioReport:
    def __init__(self, datos: pd.DataFrame, municipio, titulo_reporte):
        self.datos = datos
        self.municipio = municipio
        self.titulo = titulo_reporte
        self.datos["Clasificacion"] = self.datos["PorcentajeMora"].apply(
            self.clasificar_mora)

    def generar_pdf(self, ruta_salida="mora_vs_ingresos_gral.pdf"):
        doc = SimpleDocTemplate(ruta_salida, pagesize=landscape(letter))
        elementos = []
        estilos = getSampleStyleSheet()
        estilo_encabezado = columa_style()

        titulo = Paragraph(
            f"{self.municipio['NombreMuni']} - {self.municipio['NombreDepto']}", estilos["Title"],)
        titulo_rpt = Paragraph(
            f"REPORTE DE MORA {self.titulo}", estilos["Title"],)
        fecha = Paragraph(
            f"<b>Fecha de generación: </b> {datetime.now().strftime('%d/%m/%Y %H:%M')}", estilos["Normal"],)

        elementos.append(titulo)
        elementos.append(Spacer(1, 12))
        elementos.append(titulo_rpt)
        elementos.append(Spacer(1, 24))
        elementos.append(fecha)
        elementos.append(Spacer(1, 24))

        encabezados = [Paragraph("Anio", estilo_encabezado),
                       Paragraph("Codigo", estilo_encabezado),
                       Paragraph("Aldea", estilo_encabezado),
                       Paragraph("Total Facturas", estilo_encabezado),
                       Paragraph("Total Generado", estilo_encabezado),
                       Paragraph("Facturas Pagadas", estilo_encabezado),
                       Paragraph("Ingresos (L)", estilo_encabezado),
                       Paragraph("Factuas en Mora", estilo_encabezado),
                       Paragraph("Mora (L)", estilo_encabezado),
                       Paragraph("Porcenta Ingresos", estilo_encabezado),
                       Paragraph("Porcentaje Mora", estilo_encabezado),
                       ]
        filas = [encabezados]

        total_mora = 0
        total_ingresos = 0
        total_generado = 0

        total_fact_ingresos = 0
        total_fact_generado = 0
        total_fact_mora = 0
        total_porcent_mora = 0
        total_porcent_fact_mora = 0

        aldeas = []
        porcentajes = []

        for item in self.datos.iterrows():

            facturas = (item[1]["totalFacturas"]
                        ) if item[1]["totalFacturas"] else 0
            totalGenrado = (item[1]["TotalGenerado"]
                            ) if item[1]["TotalGenerado"] else 0
            facturas_pagadas = (
                item[1]["ingresosTotalFacturas"]) if item[1]["ingresosTotalFacturas"] else 0
            ingresos = (item[1]["Ingresos"]) if item[1]["Ingresos"] else 0
            facturas_mora = (item[1]["MoraTotalFacturas"]
                             ) if item[1]["MoraTotalFacturas"] else 0
            mora = (item[1]["MoraTotal"]) if item[1]["MoraTotal"] else 0

            porcentaje_ingresos = (
                item[1]["PorcentajeIngresos"]) if item[1]["PorcentajeIngresos"] else 0
            porcentaje_mora = (item[1]["PorcentajeMora"]
                               ) if item[1]["PorcentajeMora"] else 0

            filas.append([(item[1]["anio"]), (item[1]["CodAldea"]), item[1]["NombreAldea"].strip(), f"{facturas:,.2f}", f"{totalGenrado:,.2f}", f"{facturas_pagadas:,.2f}",
                         f"{ingresos:,.2f}", f"{facturas_mora:,.2f}", f"{mora:,.2f}", f"{porcentaje_ingresos:,.2f}", f"{porcentaje_mora:,.2f}"])

            total_mora += mora
            total_ingresos += ingresos
            total_generado += totalGenrado

            total_fact_mora += facturas_mora
            total_fact_ingresos += facturas_pagadas
            total_fact_generado += facturas

            aldeas.append(item[1]["NombreAldea"].strip())
            porcentajes.append(item[1]["PorcentajeMora"])

        total_percent_mora = (total_mora * 100)/total_generado
        total_percent_pagado = (total_ingresos * 100)/total_generado
        filas.append(["", "", "TOTAL", f"{total_fact_generado:,.2f}", f"{total_generado:,.2f}", f"{total_fact_ingresos:,.2f}", f"{total_ingresos:,.2f}",
                     f"{total_fact_mora:,.2f}", f"{total_mora:,.2f}", f"{total_percent_pagado:,.2f}", f"{total_percent_mora:,.2f}"])

        plt.figure(figsize=(15, 6))
        plt.bar(aldeas, porcentajes)
        plt.xticks(rotation=90, ha='right', fontsize=12)
        plt.title("Porcentaje de Mora por Aldea",
                  fontsize=16, fontweight="bold")
        plt.ylabel("Porcentaje de Mora (%)", fontsize=12)
        plt.xlabel("Aldea", fontsize=12)

        plt.tight_layout()
        plt.savefig("grafico_mora.png")
        plt.close()

        tabla = Table(filas, colWidths=[
            50,  60, 100, 80, 80, 80, 80, 80, 80, 60, 60])
        tabla.setStyle(
            TableStyle(
                [
                    # Encabezado
                    ("BACKGROUND", (0, 0), (-1, 0), colors.HexColor("#4F81BD")),
                    ("TEXTCOLOR", (0, 0), (-1, 0), colors.white),
                    ("FONTNAME", (0, 0), (-1, 0), "Helvetica-Bold"),
                    ("BOTTOMPADDING", (0, 0), (-1, 0), 10),

                    # Bordes y fondo general
                    ("GRID", (0, 0), (-1, -1), 0.5, colors.grey),
                    ("BACKGROUND", (0, 1), (-1, -1), colors.whitesmoke),
                    ("BACKGROUND", (-1, -1), (-1, -1), colors.lightgrey),

                    # 🔹 Alineaciones personalizadas por columna
                    # Primera columna centrada
                    ("ALIGN", (0, 0), (0, -1), "CENTER"),
                    # Segunda columna a la izquierda
                    ("ALIGN", (1, 0), (1, -1), "LEFT"),
                    # Tercera columna a la derecha
                    ("ALIGN", (2, 0), (-1, -1), "RIGHT"),
                ]
            )
        )
        texto = Paragraph(
            "Nota aclaratoria: La información de mora, ingresos y generación tributaria presentada en este reporte se basa exclusivamente en la facturación disponible en el sistema SAFT al momento de su elaboración. La ausencia de facturación actualizada o completa puede generar discrepancias en los valores mostrados, en especial en los impuestos de Bienes Inmuebles, donde puede realizar la carga masiva de facturas de las propiedades debidamente registradas.", estilos["Normal"])
        elementos.append(Image("grafico_mora.png", width=500, height=300))
        elementos.append(tabla)
        elementos.append(Spacer(1, 20))
        elementos.append(texto)
        elementos.append(Spacer(1, 20))
        footer = Paragraph(
            "<b>Generado por el sistema SAFT</b>",
            estilos["Normal"],
        )
        elementos.append(footer)

        doc.build(elementos)

    def generar_execel(self, ruta_salida="mora_vs_ingresos_gral.xlsx"):
        wb = Workbook()
        ws = wb.active
        ws.title = "Mora vs Ingresos"

        # ===============================
        # ⭐ ESTILOS
        # ===============================
        header_fill = PatternFill(
            start_color="1F4E78", end_color="1F4E78", fill_type="solid")
        header_font = Font(color="FFFFFF", bold=True)
        title_font = Font(size=16, bold=True)
        date_font = Font(size=12, italic=True)

        border = Border(
            left=Side(style="thin"), right=Side(style="thin"),
            top=Side(style="thin"), bottom=Side(style="thin")
        )

        align_center = Alignment(horizontal="center", vertical="center")
        align_right = Alignment(horizontal="right", vertical="center")
        align_left = Alignment(horizontal="left", vertical="center")

        # ===============================
        # ⭐ TITULO Y FECHA
        # ===============================
        titulo = f"{self.municipio['NombreMuni']} - {self.municipio['NombreDepto']}"
        titulo_rpt = f"REPORTE DE MORA - {self.titulo}"
        fecha = f"Fecha de elaboración: {datetime.now().strftime('%d/%m/%Y %H:%M')}"

        ws.merge_cells("A1:K1")
        ws["A1"] = titulo
        ws["A1"].alignment = align_center
        ws["A1"].font = title_font

        ws.merge_cells("A2:K2")
        ws["A2"] = titulo_rpt
        ws["A2"].alignment = align_center
        ws["A2"].font = title_font

        ws.merge_cells("A3:K3")
        ws["A3"] = fecha
        ws["A3"].alignment = align_center
        ws["A3"].font = date_font

        fila_actual = 5  # Aquí irán los encabezados
        ws.freeze_panes = "A6"
        # ===============================
        # ⭐ ENCABEZADOS
        # ===============================
        columnas = [
            "Año", "Código", "Aldea", "Total Facturas", "Total Generado",
            "Facturas Pagadas", "Ingresos (L)", "Facturas Mora", "Mora (L)",
            "% Ingresos", "% Mora"
        ]

        col_value = 0
        for col in range(1, len(columnas) + 1):
            c = ws.cell(row=fila_actual, column=col, value=columnas[col_value])
            c.fill = header_fill
            c.font = header_font
            c.alignment = align_center
            c.border = border
            col_value += 1

        fila_actual += 1

        # ===============================
        # ⭐ LLENADO DE DATOS
        # ===============================
        total_mora = 0
        total_ingresos = 0
        total_generado = 0
        total_facturas = 0
        total_facturas_pagadas = 0
        total_facturas_mora = 0

        datos_grafico = []

        for index, item in self.datos.iterrows():

            facturas = item["totalFacturas"] or 0
            totalGen = item["TotalGenerado"] or 0
            fact_pagadas = item["ingresosTotalFacturas"] or 0
            ingresos = item["Ingresos"] or 0
            fact_mora = item["MoraTotalFacturas"] or 0
            mora = item["MoraTotal"] or 0
            porcentaje_ing = item["PorcentajeIngresos"] or 0
            porcentaje_mora = item["PorcentajeMora"] or 0

            ws.append([
                item["anio"],
                item["CodAldea"].strip(),
                item["NombreAldea"].strip(),
                facturas,
                totalGen,
                fact_pagadas,
                round(ingresos, 2),
                fact_mora,
                round(mora, 2),
                round(porcentaje_ing / 100, 2),
                round(porcentaje_mora / 100, 2),
            ])

            datos_grafico.append((item["anio"], porcentaje_mora))

            # Sumatorias
            total_facturas += facturas
            total_generado += totalGen
            total_facturas_pagadas += fact_pagadas
            total_ingresos += ingresos
            total_facturas_mora += fact_mora
            total_mora += mora

            fila_actual += 1

        # ===============================
        # ⭐ FILA DE TOTALES
        # ===============================
        porcentaje_total_mora = (
            total_mora * 100 / total_generado) if total_generado else 0
        porcentaje_total_ing = (total_ingresos * 100 /
                                total_generado) if total_generado else 0

        ws.append(["Total",
                   "",
                   "",
                   total_facturas,
                   total_generado,
                   total_facturas_pagadas,
                   round(total_ingresos, 2),
                   total_facturas_mora,
                   round(total_mora, 2),
                   round(porcentaje_total_ing / 100, 2),
                   round(porcentaje_total_mora / 100, 2),
                   ])

        # ===============================
        # ⭐ FORMATO DE CELDAS Y MONEDA
        # ===============================
        ultima_fila = ws.max_row
        fila = 5

        for row in ws.iter_rows(min_row=5, max_row=ultima_fila, max_col=11):
            for cell in row:
                cell.border = border
                fila += 1
                porcentaje = ws.cell(row=fila, column=10).value

                if cell.column_letter in ["E", "G", "I"]:  # MONEDA
                    cell.number_format = '"L" #,##0.00'

                if cell.column_letter in ["D", "F", "H"]:  # Cantidades
                    cell.number_format = "#,##0"

                if cell.column_letter in ["J", "K"]:  # PORCENTAJES
                    cell.number_format = "0.00%"

                if cell.column_letter in ["A", "B", "C"]:
                    cell.alignment = align_center
        for row in range(6, ultima_fila + 1):
            cell = ws[f"K{row}"]          # Celda actual en columna J
            # Valor numérico (entre 0 y 1 si es %)
            porcentaje = cell.value

            if porcentaje is None:
                continue

            # Si el porcentaje está en formato porcentaje (0.25 = 25%)
            porcentaje_real = porcentaje * 100

            # Mapa de calor
            if porcentaje_real >= 91:
                color = "FF0000"  # ROJO fuerte (Crítica)
            elif porcentaje_real >= 70:
                color = "FF7F7F"  # Rojo suave (Muy alta)
            elif porcentaje_real >= 40:
                color = "FFC000"  # Naranja (Alta)
            elif porcentaje_real >= 20:
                color = "FFFF00"  # Amarillo (Media)
            elif porcentaje_real >= 10:
                color = "92D050"  # Verde suave (Baja)
            else:
                color = "00B050"  # Verde fuerte (Excelente)

            # Aplicar color
            cell.fill = PatternFill(
                start_color=color, end_color=color, fill_type="solid")
        # ===============================
        # ⭐ AJUSTE AUTOMÁTICO DE COLUMNAS
        # ===============================
        for col in ws.columns:
            max_length = 0
            column = col[0].column  # número de columna
            column_letter = get_column_letter(column)

            for cell in col[5:]:
                try:
                    if cell.value:
                        length = len(str(cell.value))
                        if length > max_length:
                            max_length = length
                except:
                    pass
            adjusted_width = max_length + 2  # un pequeño margen
            ws.column_dimensions[column_letter].width = adjusted_width
        # ===============================
        # ⭐ GRÁFICO INCRUSTADO
        # ===============================
        fila_inicio = 6              # fila del encabezado
        fila_datos = fila_inicio
        fila = 5
        fila_fin = fila_datos + len(datos_grafico)   # 7 + 15 - 1 = 21

        chart = BarChart()
        chart.type = "col"
        chart.title = "Porcentaje de Mora de Aldea por año"
        chart.y_axis.title = "Porcentaje (%)"
        chart.x_axis.title = "AÑO"

        # %Mora está en la columna K (11)
        data = Reference(ws, min_col=11, min_row=fila, max_row=fila_fin)

        # Año está en la columna A (1)
        categorias = Reference(
            ws, min_col=1, min_row=fila_inicio, max_row=fila_fin)

        chart.add_data(data, titles_from_data=True)
        chart.set_categories(categorias)

        ws.add_chart(chart, "M5")

        # ===============================
        # ⭐ GUARDAR
        # ===============================
        wb.save(ruta_salida)
        return ruta_salida

    def clasificar_mora(self, porcentaje):
        if porcentaje:
            if porcentaje >= 91:
                return "Crítica"
            elif porcentaje >= 70:
                return "Muy Alta"
            elif porcentaje >= 40:
                return "Alta"
            elif porcentaje >= 20:
                return "Media"
            elif porcentaje >= 10:
                return "Baja"
            else:
                return "Excelente"
        return "No Calculado"
