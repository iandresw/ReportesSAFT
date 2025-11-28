
from datetime import datetime
from openpyxl.utils import get_column_letter
from openpyxl import Workbook
from openpyxl.chart import BarChart, Reference
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side

import pandas as pd
from ui.ui_style_table import columa_style


class AnalisisIngresosReport:
    def __init__(self, datos: pd.DataFrame, municipio, titulo_reporte, anio):
        self.datos = datos
        self.municipio = municipio
        self.anio_act = anio
        self.anio_ant = int(anio) - 1
        self.titulo = titulo_reporte

    def generar_execel(self, ruta_salida="mora_vs_ingresos_gral.xlsx"):
        wb = Workbook()
        ws = wb.active
        ws.title = f"{self.anio_act} vs {self.anio_ant}"

        # ===============================
        # ⭐ ESTILOS
        # ===============================
        header_fill = PatternFill(
            start_color="1F4E78", end_color="1F4E78", fill_type="solid")
        header_font = Font(color="FFFFFF", bold=True)
        title_font = Font(size=16, bold=True)
        date_font = Font(size=12, italic=True)

        border = Border(
            left=Side(style="thin"), right=Side(style="thin"),
            top=Side(style="thin"), bottom=Side(style="thin")
        )

        align_center = Alignment(horizontal="center", vertical="center")
        align_right = Alignment(horizontal="right", vertical="center")
        align_left = Alignment(horizontal="left", vertical="center")

        # ===============================
        # ⭐ TITULO Y FECHA
        # ===============================
        titulo = f"{self.municipio['NombreMuni']} - {self.municipio['NombreDepto']}"
        titulo_rpt = f"ANALISIS DE INGRESOS - {self.anio_ant} vs {self.anio_act}"
        fecha = f"Fecha de elaboración: {datetime.now().strftime('%d/%m/%Y %H:%M')}"

        ws.merge_cells("A1:AB1")
        ws["A1"] = titulo
        ws["A1"].alignment = align_center
        ws["A1"].font = title_font

        ws.merge_cells("A2:AB2")
        ws["A2"] = titulo_rpt
        ws["A2"].alignment = align_center
        ws["A2"].font = title_font

        ws.merge_cells("A3:AB3")
        ws["A3"] = fecha
        ws["A3"].alignment = align_center
        ws["A3"].font = date_font

        fila_actual = 5  # Aquí irán los encabezados
        ws.freeze_panes = "A6"
        # ===============================
        # ⭐ ENCABEZADOS
        # ===============================
        columnas = [
            "Código", "Nombre",
            f"Ingresos Enero ({self.anio_ant})(C)", f"Ingresos Enero ({self.anio_act})(D)",
            f"Ingreso Diferencia E = (D - C)",
            f"Ingresos Febrero ({self.anio_ant})(C)", f"Ingresos Febrero ({self.anio_act})(D)",
            f"Ingreso Diferencia E = (D - C)",
            f"Ingresos Marzo ({self.anio_ant})(C)", f"Ingresos Marzo ({self.anio_act})(D)",
            f"Ingreso Diferencia E = (D - C)",
            f"Ingresos Abril ({self.anio_ant})(C)", f"Ingresos Abril ({self.anio_act})(D)",
            f"Ingreso Diferencia E = (D - C)",
            f"Ingresos Mayo ({self.anio_ant})(C)", f"Ingresos Mayo ({self.anio_act})(D)",
            f"Ingreso Diferencia E = (D - C)",
            f"Ingresos Junio ({self.anio_ant})(C)", f"Ingresos Junio ({self.anio_act})(D)",
            f"Ingreso Diferencia E = (D - C)",
            f"Ingresos Julio ({self.anio_ant})(C)", f"Ingresos Julio ({self.anio_act})(D)",
            f"Ingreso Diferencia E = (D - C)",
            f"Ingresos Agosto ({self.anio_ant})(C)", f"Ingresos Agosto ({self.anio_act})(D)",
            f"Ingreso Diferencia E = (D - C)",
            f"Ingresos Septiembre ({self.anio_ant})(C)", f"Ingresos Septiembre ({self.anio_act})(D)",
            f"Ingreso Diferencia E = (D - C)",
            f"Ingresos Octubre ({self.anio_ant})(C)", f"Ingresos Octubre ({self.anio_act})(D)",
            f"Ingreso Diferencia E = (D - C)",
            f"Ingresos Noviembre ({self.anio_ant})(C)", f"Ingresos Noviembre ({self.anio_act})(D)",
            f"Ingreso Diferencia E = (D - C)",
            f"Ingresos Diciembre ({self.anio_ant})(C)", f"Ingresos Diciembre ({self.anio_act})(D)",
            f"Ingreso Diferencia E = (D - C)",
            f"Total ({self.anio_ant})",
            f"Total ({self.anio_act})"
            f"Ingreso Diferencia E = (D - C)",
        ]

        col_value = 0
        for col in range(1, len(columnas) + 1):
            c = ws.cell(row=fila_actual, column=col, value=columnas[col_value])
            c.fill = header_fill
            c.font = header_font
            c.alignment = align_center
            c.border = border
            col_value += 1

        fila_actual += 1

        # ===============================
        # ⭐ LLENADO DE DATOS
        # ===============================
        tot_enero_01 = 0
        tot_enero_02 = 0
        tot_febrero_01 = 0
        tot_febrero_02 = 0
        tot_marzo_01 = 0
        tot_marzo_02 = 0
        tot_abril_01 = 0
        tot_abril_02 = 0
        tot_mayo_01 = 0
        tot_mayo_02 = 0
        tot_junio_01 = 0
        tot_junio_02 = 0
        tot_julio_01 = 0
        tot_julio_02 = 0
        tot_agosto_01 = 0
        tot_agosto_02 = 0
        tot_septiembre_01 = 0
        tot_septiembre_02 = 0
        tot_octubre_01 = 0
        tot_octubre_02 = 0
        tot_noviembre_01 = 0
        tot_noviembre_02 = 0
        tot_diciembre_01 = 0
        tot_diciembre_02 = 0
        tot_total_ant = 0
        tot_total_act = 0

        for index, item in self.datos.iterrows():

            enero_01 = item["enero_ant"] or 0
            enero_02 = item["enero_act"] or 0
            enero_03 = enero_02 - enero_01

            febrero_01 = item["febrero_ant"] or 0
            febrero_02 = item["febrero_act"] or 0
            febrero_03 = febrero_02-febrero_01

            marzo_01 = item["marzo_ant"] or 0
            marzo_02 = item["marzo_act"] or 0
            marzo_03 = marzo_02-marzo_01

            abril_01 = item["abril_ant"] or 0
            abril_02 = item["abril_act"] or 0
            abril_03 = abril_02-abril_01

            mayo_01 = item["mayo_ant"] or 0
            mayo_02 = item["mayo_act"] or 0
            mayo_03 = mayo_02-mayo_01

            junio_01 = item["junio_ant"] or 0
            junio_02 = item["junio_act"] or 0
            junio_03 = junio_02-junio_01

            julio_01 = item["julio_ant"] or 0
            julio_02 = item["julio_act"] or 0
            julio_03 = julio_02-julio_01

            agosto_01 = item["agosto_ant"] or 0
            agosto_02 = item["agosto_act"] or 0
            agosto_03 = agosto_02-agosto_01

            septiembre_01 = item["septiembre_ant"] or 0
            septiembre_02 = item["septiembre_act"] or 0
            septiembre_03 = septiembre_02-septiembre_01

            octubre_01 = item["octubre_ant"] or 0
            octubre_02 = item["octubre_act"] or 0
            octubre_03 = octubre_02-octubre_01

            noviembre_01 = item["noviembre_ant"] or 0
            noviembre_02 = item["noviembre_act"] or 0
            noviembre_03 = noviembre_02-noviembre_01

            diciembre_01 = item["diciembre_ant"] or 0
            diciembre_02 = item["diciembre_act"] or 0
            diciembre_03 = diciembre_02-diciembre_01

            total_ant = enero_01 + febrero_01 + marzo_01 + abril_01 + mayo_01 + junio_01 + \
                julio_01 + agosto_01 + septiembre_01 + octubre_01 + noviembre_01 + diciembre_01
            total_act = enero_02 + febrero_02 + marzo_02 + abril_02 + mayo_02 + junio_02 + \
                julio_02 + agosto_02 + septiembre_02 + octubre_02 + noviembre_02 + diciembre_02

            total_dif = total_act - total_ant

            ws.append([
                item["cta"].strip(),
                "NombreCta",
                enero_01,
                enero_02,
                enero_03,
                febrero_01,
                febrero_02,
                febrero_03,
                marzo_01,
                marzo_02,
                marzo_03,
                abril_01,
                abril_02,
                abril_03,
                mayo_01,
                mayo_02,
                mayo_03,
                junio_01,
                junio_02,
                junio_03,
                julio_01,
                julio_02,
                julio_03,
                agosto_01,
                agosto_02,
                agosto_03,
                septiembre_01,
                septiembre_02,
                septiembre_03,
                octubre_01,
                octubre_02,
                octubre_03,
                noviembre_01,
                noviembre_02,
                noviembre_03,
                diciembre_01,
                diciembre_02,
                diciembre_03,
                total_ant,
                total_act,
                total_dif
            ])

            tot_enero_01 += enero_01
            tot_enero_02 += enero_02
            tot_febrero_01 += febrero_01
            tot_febrero_02 += febrero_02
            tot_marzo_01 += marzo_01
            tot_marzo_02 += marzo_02
            tot_abril_01 += abril_01
            tot_abril_02 += abril_02
            tot_mayo_01 += mayo_01
            tot_mayo_02 += mayo_02
            tot_junio_01 += junio_01
            tot_junio_02 += junio_02
            tot_julio_01 += julio_01
            tot_julio_02 += julio_02
            tot_agosto_01 += agosto_01
            tot_agosto_02 += agosto_02
            tot_septiembre_01 += septiembre_01
            tot_septiembre_02 += septiembre_02
            tot_octubre_01 += octubre_01
            tot_octubre_02 += octubre_02
            tot_noviembre_01 += noviembre_01
            tot_noviembre_02 += noviembre_02
            tot_diciembre_01 += diciembre_01
            tot_diciembre_02 += diciembre_02
            tot_total_ant += total_ant
            tot_total_act += total_act

        # ===============================
        # ⭐ FILA DE TOTALES
        # ===============================

        ws.append([
            "",
            "TOTAL",
            tot_enero_01, tot_enero_02, tot_febrero_01, tot_febrero_02, tot_marzo_01, tot_marzo_02, tot_abril_01, tot_abril_02, tot_mayo_01, tot_mayo_02, tot_junio_01, tot_junio_02, tot_julio_01, tot_julio_02, tot_agosto_01, tot_agosto_02, tot_septiembre_01, tot_septiembre_02, tot_octubre_01, tot_octubre_02, tot_noviembre_01, tot_noviembre_02, tot_diciembre_01, tot_diciembre_02, tot_total_ant, tot_total_act
        ])

        # ===============================
        # ⭐ FORMATO DE CELDAS Y MONEDA
        # ===============================
        ultima_fila = ws.max_row
        fila = 5

        for row in ws.iter_rows(min_row=5, max_row=ultima_fila, max_col=42):
            for cell in row:
                cell.border = border
                fila += 1
                porcentaje = ws.cell(row=fila, column=10).value

                if cell.column_letter in ["A", "B"]:  # Total facturas
                    cell.alignment = align_right
                    cell.number_format = "#,##0"

                if cell.column_letter in ["C", "D", "E", "F", "G", "H", "I", "J", "K", "L", "M", "N", "O", "P", "Q", "R", "S", "T", "U", "V", "W", "X", "Y", "Z", "AB"]:  # MONEDA
                    cell.number_format = '"L" #,##0.00'

        # ===============================
        # ⭐ AJUSTE AUTOMÁTICO DE COLUMNAS
        # ===============================
        for col in ws.columns:
            max_length = 0
            column = col[0].column  # número de columna
            column_letter = get_column_letter(column)

            for cell in col[5:]:
                try:
                    if cell.value:
                        length = len(str(cell.value))
                        if length > max_length:
                            max_length = length
                except:
                    pass
            adjusted_width = max_length + 2  # un pequeño margen
            ws.column_dimensions[column_letter].width = adjusted_width

        for row in range(6, ultima_fila + 1):
            cell = ws[f"E{row}"]          # Celda actual en columna J

            valor = cell.value
            if valor is None:
                continue

            if valor < 0:
                color = "FF7F7F"
            elif valor >= 0:
                color = "00B050"

            # Aplicar color
            cell.fill = PatternFill(
                start_color=color, end_color=color, fill_type="solid")

        # ===============================
        # ⭐ GRÁFICO INCRUSTADO
        # ===============================

        chart = BarChart()
        chart.type = "col"
        chart.title = "Porcentaje de Mora por Aldea"
        chart.y_axis.title = "Porcentaje (%)"
        chart.x_axis.title = "Aldeas"

        data = Reference(ws, min_col=10, min_row=6,
                         max_row=len(self.datos)+1)
        categorias = Reference(ws, min_col=2, min_row=6,
                               max_row=len(self.datos)+1)
        chart.add_data(data, titles_from_data=True)
        chart.set_categories(categorias)

        ws.add_chart(chart, "L28")

        # ===============================
        # ⭐ GUARDAR
        # ===============================
        wb.save(ruta_salida)
        return ruta_salida

    def clasificar_mora(self, porcentaje):
        if porcentaje:
            if porcentaje >= 91:
                return "Crítica"
            elif porcentaje >= 70:
                return "Muy Alta"
            elif porcentaje >= 40:
                return "Alta"
            elif porcentaje >= 20:
                return "Media"
            elif porcentaje >= 10:
                return "Baja"
            else:
                return "Excelente"
        return "No Calculado"
